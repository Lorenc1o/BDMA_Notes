import openpyxl
import random
import string
import heapq

def create_voting_profiles():
    # Create a new workbook
    wb = openpyxl.Workbook()
    # Select the active worksheet
    ws = wb.active
    # Define the voting profiles
    profiles = [
        [5, ['a', 'b', 'c', 'd']],
        [4, ['a', 'c', 'b', 'd']],
        [2, ['d', 'b', 'a', 'c']],
        [6, ['d', 'b', 'c', 'a']],
        [8, ['c', 'b', 'a', 'd']],
        [2, ['d', 'c', 'b', 'a']]
    ]
    # Write the voting profiles to the worksheet
    for profile in profiles:
        for _ in range(profile[0]):
            ws.append(profile[1])
    # Save the workbook, overwriting the previous file
    wb.save('data.xlsx')

def read_voting_profiles(datafile):
    # Open the workbook
    wb = openpyxl.load_workbook(datafile)
    # Select the active worksheet
    ws = wb.active
    # Read the voting profiles
    profiles = {}
    for row in ws.iter_rows():
        profile = ''
        for cell in row:
            profile += cell.value
        
        if profile not in profiles.keys():
            profiles[profile] = 1
        else:
            profiles[profile] += 1
    
    return profiles

def SimpleMajorityRulefor2(profiles,a = 'a', b = 'b'):
    '''
    Given a set of voting profiles, and two alternatives a and b, return the winner according to the simple majority rule
    '''
    # Initialize the votes
    votes = {a: 0, b: 0}
    # Count the votes, X gets a vote if X is preferred to Y in the profile
    try:
        for profile in profiles:
            a_idx = profile.index(a)
            b_idx = profile.index(b)
            if a_idx < b_idx:
                votes[a] += profiles[profile]
            else:
                votes[b] += profiles[profile]
        # Return the winner
        if votes[a] > votes[b]:
            return a
        else:
            return b
    except:
        print('\033[91mInvalid alternatives!\033[00m')
        return None
    
def Plurality(profiles):
    '''
    Given a set of profiles, return the winner according to the plurality rule.
    In case of a tie, resolve it by lexicographic order.
    '''
    # Initialize the votes
    votes = {}
    # Count the votes
    for profile in profiles:
        if profile[0] not in votes.keys():
            votes[profile[0]] = 1
        else:
            votes[profile[0]] += 1
    # Return the winner, in case of a tie, resolve it by lexicographic order
    max_votes = max(votes.values())
    winners = []
    for candidate in votes:
        if votes[candidate] == max_votes:
            winners.append(candidate)
    return min(winners)

def PluralityRunoff(profiles):
    '''
    Given a set of profiles, return the winner according to the plurality runoff rule.
    In case of a tie, resolve it by lexicographic order.
    '''
    # Initialize the votes for the first round
    votes = {}
    # Count the votes for the first round
    for profile in profiles:
        if profile[0] not in votes.keys():
            votes[profile[0]] = 1
        else:
            votes[profile[0]] += 1
    # Second round: Find the two candidates with the most votes
    first_round = [max(votes, key=lambda x: (votes[x], x))]

    # If in the first round there is a candidate with more than half of the votes, return it
    if votes[first_round[0]] > sum(votes.values())/2:
        return first_round[0]
    
    votes.pop(first_round[0])
    first_round.append(max(votes, key=lambda x: (votes[x], x)))
    # Return the winner of the second round
    return SimpleMajorityRulefor2(profiles, first_round[0], first_round[1])

def CondorcetVoting(profiles):
    '''
    Given a set of profiles, return the winner according to the Condorcet voting rule.
    If there is no Condorcet winner, return None.
    '''
    # Get the alternatives
    alternatives = []
    # It's enough to check the first profile
    key_profile = list(profiles.keys())[0]
    for alternative in key_profile:
        alternatives.append(alternative)

    # Initialize the votes
    votes = {}
    # Count the votes
    for i in range(len(alternatives)-1):
        for j in range(i+1, len(alternatives)):
            winner = SimpleMajorityRulefor2(profiles, alternatives[i], alternatives[j])
            if winner not in votes.keys():
                votes[winner] = 1
            else:
                votes[winner] += 1
    # Return the winner, if there is no Condorcet winner, return None
    if max(votes.values()) == len(alternatives) - 1:
        return max(votes, key=lambda x: (votes[x], x))
    else:
        return None

def BordaVoting(profiles):
    '''
    Given a set of profiles, return the winner according to the Borda voting rule.
    In case of a tie, resolve it by lexicographic order.
    '''
    # Initialize the votes
    votes = {}
    # Count the votes
    for profile in profiles:
        for i in range(len(profile)):
            if profile[i] not in votes.keys():
                votes[profile[i]] = (i+1)*profiles[profile]
            else:
                votes[profile[i]] += (i+1)*profiles[profile]
    # Return the winner, in case of a tie, resolve it by lexicographic order
    return min(votes, key=lambda x: (votes[x], x))

###################################
# Here, we perform the generation #
# using the theorem               #
###################################

def use_theorem(n=40, m=6):
    solution = {}
    # First profile: abc...
    profile = ''
    for i in range(m):
        profile += string.ascii_lowercase[i]

    # Second profile: same, change the first two alternatives
    profile2 = list(profile)
    profile2[0], profile2[1] = profile2[1], profile2[0]
    profile2 = ''.join(profile2)
    
    if n % 2 == 0:
        solution[profile] = n//2+1
        solution[profile2] = n//2-1
    else:
        solution[profile] = n//2+1
        solution[profile2] = n//2

    return solution

###################################
# Here, we perform the generation #
# randomly                        #
###################################

def generateRandomProfile(m=6):
    '''
    Generate a random profile with m alternatives
    '''
    alternatives = list(string.ascii_lowercase)[:m]
    random.shuffle(alternatives)
    # Return the profile as a string
    return ''.join(alternatives)

def generateRandomProfiles(n=40, m=6, thr=0.25):
    '''
    Generate n random profiles with m alternatives
    '''
    gen_profiles = {}
    profile = None
    thr = thr / min(20,n)
    for _ in range(n):
        rnd = random.random()
        if rnd < thr or profile == None:
            profile = generateRandomProfile(m)
        if profile not in gen_profiles.keys():
            gen_profiles[profile] = 1
        else:
            gen_profiles[profile] += 1
    return gen_profiles

def check_preferences(profiles):
    '''
    Check if at least 10% of voters have different preferences
    If the maximum number of voters with the same preferences is greater than 90% of the total number of voters, return False
    '''
    max_votes = max(profiles.values())
    if max_votes > 0.9*sum(profiles.values()):
        return False
    else:
        return True
    
def check_candidates(profiles):
    '''
    Check it no more than 70% of voters have the same best candidate
    If the number of voters with the best candidate is greater than 70% of the total number of voters, return False
    '''
    best_candidates = {}
    for profile in profiles:
        if profile[0] not in best_candidates.keys():
            best_candidates[profile[0]] = profiles[profile]
        else:
            best_candidates[profile[0]] += profiles[profile]
    max_votes = max(best_candidates.values())
    if max_votes > 0.7*sum(profiles.values()):
        return False
    else:
        return True
    
def validate(profiles):
    '''
    Validate the profiles
    '''
    if check_preferences(profiles) and check_candidates(profiles):
        return True
    else:
        return False
    
def generateValidProfiles(n=40, m=6, thr=0.25):
    '''
    Generate n valid profiles with m alternatives
    '''
    profiles = generateRandomProfiles(n, m, thr)
    while not validate(profiles):
        profiles = generateRandomProfiles(n, m, thr)
    return profiles

def allWinner(n=40, m=6):
    '''
    Generate a profile with n voters and m candidates in which the winner wins all the types of voting rules
    '''
    quantity = 0
    while True:
        profiles = generateValidProfiles(n, m)
        quantity += 1
        if Plurality(profiles) == PluralityRunoff(profiles) == CondorcetVoting(profiles) == BordaVoting(profiles):
            return profiles, quantity
        
def eachWinner(n=40, m=6):
    '''
    Generate a profile with n voters and m candidates in which a different candidate wins each type of voting rule
    '''
    quantity = 0
    while True:
        profiles = generateValidProfiles(n, m)
        quantity += 1
        if CondorcetVoting(profiles) is None:
            continue

        winners = set([Plurality(profiles), PluralityRunoff(profiles), CondorcetVoting(profiles), BordaVoting(profiles)])

        if len(winners) == 4:
            return profiles, quantity
        
###################################
# Here, we perform the generation #
# using genetic algorithms        #
###################################
# 1. Representation: Each voter's preference is a list of candidates, and each solution is a list of all voters' preferences.
def generate_random_solution(num_voters, num_candidates, mode='all'):
    if mode == 'all':
        return generateRandomProfiles(num_voters, num_candidates)
    elif mode == 'diverse':
        return generateValidProfiles(num_voters, num_candidates, thr=1)

def calculate_fitness(solution, n=40, m=6, mode='all'):

    winners = set([Plurality(solution), PluralityRunoff(solution), BordaVoting(solution)])

    condorcet = CondorcetVoting(solution)

    if condorcet is not None:
        winners.add(condorcet)

    num_winners = len(winners)

    # Initialize variables to hold information about the current solution.
    candidate_wins = {candidate: 0 for candidate in list(string.ascii_lowercase)[:m]}
    profile_counts = {}
    top_candidate_counts = {candidate: 0 for candidate in list(string.ascii_lowercase)[:m]}

    diverse = 1 if check_preferences(solution) else 0

    diverse_candidates = 1 if check_candidates(solution) else 0

    if mode=='all':
        justone = 1 if num_winners == 1 else 0
        return justone*3 + diverse*2 + diverse_candidates
    elif mode=='diverse':
        return num_winners*2 + diverse + diverse_candidates

def select_parents(population, fitnesses):
    # Calculate the total fitness of all individuals in the population.
    total_fitness = sum(fitnesses)
    
    # If all solutions are equally unfit, random selection can be performed
    if total_fitness == 0:
        return random.choice(population), random.choice(population)

    # Normalize individual fitnesses between 0 and 1.
    normalized_fitnesses = [f / total_fitness for f in fitnesses]

    # Calculate the cumulative probability for each individual.
    cumulative_prob = [sum(normalized_fitnesses[:i+1]) for i in range(len(normalized_fitnesses))]

    # Select two parents via roulette wheel selection.
    parents = []
    for _ in range(2):  # Doing this twice to select two parents
        # Generate a random number and find the first individual that has a higher cumulative probability.
        rand = random.random()
        for i in range(len(cumulative_prob)):
            if rand <= cumulative_prob[i]:
                parents.append(population[i])
                break

    return parents[0], parents[1]  # return the two selected parents

# 4. Crossover: Combine two solutions to make a new solution.
def crossover(parent1, parent2, n, crossover_rate=1):
    '''
    Combines two parent voting profiles into a single child profile, preferring profiles with higher counts.
    parent1, parent2: Dictionaries representing counts of different voting profiles.
    n: The total number of voters to be represented in the child.
    crossover_rate: The probability that crossover will occur.
    '''
    # If the parents are the same, just return one of them.
    if parent1 == parent2:
        return parent1

    # If the random value exceeds the crossover rate, just pick one parent entirely at random for the next generation.
    if random.random() >= crossover_rate:
        return parent1 if random.random() < 0.5 else parent2

    # Merge the profiles from both parents, summing the counts for each profile.
    all_profiles = {}

    for profile, count in parent1.items():
        if profile not in all_profiles.keys():
            all_profiles[profile] = count
        else:
            all_profiles[profile] += count

    for profile, count in parent2.items():
        if profile not in all_profiles.keys():
            all_profiles[profile] = count
        else:
            all_profiles[profile] += count

    # Now we add profiles until we reach the total number of voters.
    child = {}
    while sum(child.values()) < n:
        for profile, count in all_profiles.items():
            if sum(child.values()) + 1 <= n:
                if profile not in child.keys():
                    child[profile] = 1
                    all_profiles[profile] -= 1
                else:
                    child[profile] += 1
                    all_profiles[profile] -= 1
                if all_profiles[profile] == 0:
                    all_profiles.pop(profile)
            else:
                break

    return child

# 5. Mutation: Randomly alter the solutions.
def mutate(solution, mutation_rate=1, mode='all'):
    # If the solution only has one profile, we generate a new profile and split the votes between the two profiles.
    if len(solution) == 1:
        profile = generateRandomProfile(len(list(solution.keys())[0]))
        solution[profile] = solution.pop(list(solution.keys())[0]) // 2
        solution[list(solution.keys())[0]] -= solution[list(solution.keys())[0]] // 2
        return

    if len(solution) == 2:
        profile = generateRandomProfile(len(list(solution.keys())[0]))
        solution[profile] = solution.pop(list(solution.keys())[0]) // 2
        solution[list(solution.keys())[0]] -= solution[list(solution.keys())[0]] // 2
        return
    
    # If the random value exceeds the mutation rate, we don't mutate this solution.
    if random.random() < mutation_rate:
        # Remove the least common profile from the solution and add the votes to the most common profile.
        # 1- Find the least common profile
        min_count = min(solution.values())
        min_profile = None
        for profile, count in solution.items():
            if count == min_count:
                min_profile = profile
                break

        # 2- Find the most common profile
        max_count = max(solution.values())
        max_profile = None
        for profile, count in solution.items():
            if count == max_count:
                max_profile = profile
                break

        # 3- Remove the least common profile and add the votes to the most common profile
        if min_profile != max_profile:
            solution[max_profile] += solution[min_profile]
            solution.pop(min_profile)

        if mode == 'diverse' and len(solution) < 4:
            # If the problem is to have different winners, for each voting rule,
            # We take the most common profile and divide it into three profiles with the same number of votes
            # Example:
            #  - Most common profile: 'abcd', 10 votes
            #  - We divide it into 'abdc', 'bacd' and 'cbad', each with 3 votes (the remaining vote goes to the original profile)
            max_profile_count = solution[max_profile]
            max_profile_variant = list(max_profile)
            max_profile_variant[0], max_profile_variant[1] = max_profile_variant[1], max_profile_variant[0]
            max_profile_variant = ''.join(max_profile_variant)
            
            if max_profile_variant not in solution.keys():
                solution[max_profile_variant] = max_profile_count // 3
            else:
                solution[max_profile_variant] += max_profile_count // 3

            solution[max_profile] -= max_profile_count // 3
            max_profile_variant = list(max_profile)
            max_profile_variant[0], max_profile_variant[2] = max_profile_variant[2], max_profile_variant[0]
            max_profile_variant = ''.join(max_profile_variant)

            if max_profile_variant not in solution.keys():
                solution[max_profile_variant] = max_profile_count // 3
            else:
                solution[max_profile_variant] += max_profile_count // 3

            solution[max_profile] -= max_profile_count // 3

        elif mode == 'diverse' and len(solution) > 8:
            while len(solution) > 8:
                # We add the 8+k most common profile to the k-th most common profile
                k = len(solution) - 8
                kth_profile = heapq.nlargest(k, solution, key=solution.get)[-1]
                kth_least_profile = heapq.nsmallest(k, solution, key=solution.get)[-1]
                solution[kth_profile] += solution[kth_least_profile]
                solution.pop(kth_least_profile)

# Bringing it all together
def run_genetic_algorithm(n=40, m=6, population_size=100, max_generations=1000, mode='all'):
    # Generate the initial population
    population = [generate_random_solution(n, m, mode) for _ in range(population_size)]

    for gen in range(max_generations):
        # Calculate fitness for the population
        fitnesses = [calculate_fitness(solution, n, m, mode) for solution in population]

        # Check for solution that satisfies all conditions, and could be considered as the 'winner'
        for solution, fitness in zip(population, fitnesses):
            if mode == 'diverse' and fitness >= 10: # The fitness score is the sum of the weights, so a score of 10 means all conditions are met.
                return solution, (gen+1)*population_size
            elif mode == 'all' and fitness >= 6: # The fitness score is the sum of the weights, so a score of 6 means all conditions are met.
                return solution, (gen+1)*population_size

        # Create the next generation
        new_population = []
        while len(new_population) < population_size:
            parent1, parent2 = select_parents(population, fitnesses)
            try:
                child = crossover(parent1, parent2, n)
            except:
                child = generate_random_solution(n, m, mode)
                continue

            mutate(child, mode=mode)
            new_population.append(child)

        population = new_population

    print("No solution met all conditions in {} generations.".format(max_generations))
    return None, (max_generations+1)*population_size

if __name__ == '__main__':
    datafile = input('Enter the path to the data file (default: data.xlsx) or input "create" to create a new data file: ')
    if datafile == '':
        datafile = 'data.xlsx'
    if datafile == 'create':
        create_voting_profiles()
    else:
        profiles = read_voting_profiles(datafile)
    
    option = ''

    while option != '0':
        option = ''
        option = input('\033[93mEnter the action you want to perform:\n'+
                    '\t1. Show the voting profiles\n'+
                    '\t2. Simple majority rule for two alternatives\n'+
                    '\t3. Plurality rule\n'+
                    '\t4. Plurality runoff rule\n'+
                    '\t5. Condorcet voting rule\n'+
                    '\t6. Borda voting rule\n'+
                    '\t7. Generate a profile that wins all (using the theorem)\n'+
                    '\t8. Generate a profile that wins all (randomly)\n'+
                    '\t9. Generate a profile in which a different candidate wins each type of voting rule\n'+
                    '\t10. Generate a profile that wins all (genetic algorithm)\n'+
                    '\t11. Generate a profile in which a different candidate wins each type of voting rule (genetic algorithm)\n'+
                    '\t12. Compare the performance of the genetic algorithm with the random generation for the problem of generating a profile that wins all\n'+
                    '\t13. Compare the performance of the genetic algorithm with the random generation for the problem of generating a profile in which a different candidate wins each type of voting rule\n'+                    
                    '\t14. Enter new data\n'
                    '\t0. Exit\n'+
                    'Your choice: \033[00m')
        if option == '1':
            for profile in profiles: # profile is formatted: 'abcd' -> a>b>c>d
                formatted_profile = ''
                for alternative in profile:
                    formatted_profile += alternative + '>'
                print("There are", profiles[profile], "with profile", formatted_profile[:-1])
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '2':
            a = input('Enter the first alternative: ')
            b = input('Enter the second alternative: ')
            winner = SimpleMajorityRulefor2(profiles, a, b)
            if winner != None:
                print('The winner is', winner)
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '3':
            print("The winner is", Plurality(profiles), "according to the plurality rule.")
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '4':
            print("The winner is", PluralityRunoff(profiles), "according to the plurality runoff rule.")
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '5':
            winner = CondorcetVoting(profiles)
            if winner != None:
                print('The winner is', winner)
            else:
                print('There is no Condorcet winner.')
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '6':
            print("The winner is", BordaVoting(profiles), "according to the Borda voting rule.")
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '7':
            n = input('Enter the number of voters(default: 40): ')
            if n == '':
                n = 40
            else:
                n = int(n)
            m = input('Enter the number of candidates(default: 6): ')
            if m == '':
                m = 6
            else:
                m = int(m)
            print('Generating the profile...')
            gen_profiles = use_theorem(n,m)
            for profile in gen_profiles: # profile is formatted: 'abcd' -> a>b>c>d
                formatted_profile = ''
                for alternative in profile:
                    formatted_profile += alternative + '>'
                print("There are", gen_profiles[profile], "with profile", formatted_profile[:-1])
            print('The winner is', Plurality(gen_profiles),'.')
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '8':
            n = input('Enter the number of voters(default: 40): ')
            if n == '':
                n = 40
            else:
                n = int(n)
            m = input('Enter the number of candidates(default: 6): ')
            if m == '':
                m = 6
            else:
                m = int(m)
            print('Generating the profile...')
            gen_profiles, quantity = allWinner(n, m)
            for profile in gen_profiles: # profile is formatted: 'abcd' -> a>b>c>d
                formatted_profile = ''
                for alternative in profile:
                    formatted_profile += alternative + '>'
                print("\tThere are", gen_profiles[profile], "with profile", formatted_profile[:-1])
            print('The winner is', Plurality(gen_profiles),'.')
            print('It took', quantity, 'tries to generate the profile.')
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '9':
            n = input('Enter the number of voters(default: 40): ')
            if n == '':
                n = 40
            else:
                n = int(n)
            m = input('Enter the number of candidates(default: 6): ')
            if m == '':
                m = 6
            else:
                m = int(m)
            print('Generating the profile...')
            gen_profiles, quantity = eachWinner(n, m)
            for profile in gen_profiles:
                formatted_profile = ''
                for alternative in profile:
                    formatted_profile += alternative + '>'
                print("\tThere are", gen_profiles[profile], "with profile", formatted_profile[:-1])
            print('The winner is', Plurality(gen_profiles),'according to the plurality rule.')
            print('The winner is', PluralityRunoff(gen_profiles),'according to the plurality runoff rule.')
            print('The winner is', CondorcetVoting(gen_profiles),'according to the Condorcet voting rule.')
            print('The winner is', BordaVoting(gen_profiles),'according to the Borda voting rule.')
            print('It took', quantity, 'tries to generate the profile.')
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '10':
            n = input('Enter the number of voters(default: 40): ')
            if n == '':
                n = 40
            else:
                n = int(n)
            m = input('Enter the number of candidates(default: 6): ')
            if m == '':
                m = 6
            else:
                m = int(m)
            print('Generating the profile...')
            gen_profiles, quantity = run_genetic_algorithm(n,m,mode='all')

            if gen_profiles is None:
                print('No solution found.')
                input('\033[96mPress enter to continue...\033[00m')
                continue

            for profile in gen_profiles: # profile is formatted: 'abcd' -> a>b>c>d
                formatted_profile = ''
                for alternative in profile:
                    formatted_profile += alternative + '>'
                print("\tThere are", gen_profiles[profile], "with profile", formatted_profile[:-1])
            print('The winner is', Plurality(gen_profiles),'.')
            print('It took', quantity, 'tries to generate the profile.')
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '11':
            n = input('Enter the number of voters(default: 40): ')
            if n == '':
                n = 40
            else:
                n = int(n)
            m = input('Enter the number of candidates(default: 6): ')
            if m == '':
                m = 6
            else:
                m = int(m)
            print('Generating the profile...')
            gen_profiles, quantity = run_genetic_algorithm(n,m,mode='diverse')

            if gen_profiles is None:
                print('No solution found.')
                input('\033[96mPress enter to continue...\033[00m')
                continue

            for profile in gen_profiles:
                formatted_profile = ''
                for alternative in profile:
                    formatted_profile += alternative + '>'
                print("\tThere are", gen_profiles[profile], "with profile", formatted_profile[:-1])
            print('The winner is', Plurality(gen_profiles),'according to the plurality rule.')
            print('The winner is', PluralityRunoff(gen_profiles),'according to the plurality runoff rule.')
            print('The winner is', CondorcetVoting(gen_profiles),'according to the Condorcet voting rule.')
            print('The winner is', BordaVoting(gen_profiles),'according to the Borda voting rule.')
            print('It took', quantity, 'tries to generate the profile.')
            input('\033[96mPress enter to continue...\033[00m')
        elif option == '12':
            # Here, we compare the performance of the genetic algorithm with the random generation for the problem of generating a profile that wins all
            # We are going to generate solutions for:
            # - n in [40,60,80,100,120,140,160,180,200]
            # - m in candidates
            # - population_size = [10, 100, 200]
            # We will generate 10 solutions for each combination of n and m, and plot the average time needed to generate the solution
            # We will do this for both the genetic algorithm and the random generation
            import time
            import matplotlib.pyplot as plt
            import numpy as np
            from tqdm import tqdm
            from statistics import mean

            voters = [50, 100, 150, 200, 300, 500, 750, 1000, 2000]
            candidates = [6, 8, 10, 12, 14, 16, 18, 20]
            pops = [10, 100, 200]

            # Genetic algorithm
            # Table with columns: n, m, population_size, time
            table = []
            for n in tqdm(voters):
                for m in candidates:
                    for population_size in pops:
                        times = []
                        for _ in range(10):
                            start = time.time()
                            _,_ = run_genetic_algorithm(n,m,population_size,mode='all')
                            end = time.time()
                            times.append(end-start)
                        table.append([n,m,population_size,mean(times)])

            # Random generation
            table_random = []
            for n in tqdm(voters):
                for m in candidates:
                    times = []
                    for _ in range(10):
                        start = time.time()
                        _,_ = allWinner(n,m)
                        end = time.time()
                        times.append(end-start)
                    table_random.append([n,m,mean(times)])

            # Plot the results
            # Population size VS avg time
            plt.figure(figsize=(12,8))
            plt.title('Population size VS avg time')
            plt.xlabel('Population size')
            plt.ylabel('Avg time (s)')
            plt.grid()
            plt.xticks(pops)
            avg_times = []
            for population_size in pops:
                avg_times.append(mean([row[3] for row in table if row[2] == population_size]))
            plt.plot(pops, avg_times, marker='o')
            plt.savefig('population_size_all.png')

            # n VS avg time
            plt.figure(figsize=(12,8))
            plt.title('number of voters VS avg time')
            plt.xlabel('number of voters')
            plt.ylabel('Avg time (s)')
            plt.grid()
            plt.xticks(voters)
            avg_times_10 = []
            avg_times_100 = []
            avg_times_200 = []
            avg_times_random = []
            for n in voters:
                avg_times_10.append(mean([row[3] for row in table if row[0] == n and row[2] == 10]))
                avg_times_100.append(mean([row[3] for row in table if row[0] == n and row[2] == 100]))
                avg_times_200.append(mean([row[3] for row in table if row[0] == n and row[2] == 200]))
                avg_times_random.append(mean([row[2] for row in table_random if row[0] == n]))

            plt.plot(voters, avg_times_10, marker='o', label='GA: Population size = 10')
            plt.plot(voters, avg_times_100, marker='o', label='GA: Population size = 100')
            plt.plot(voters, avg_times_200, marker='o', label='GA: Population size = 200')
            plt.plot(voters, avg_times_random, marker='o', label='Random generation')
            
            plt.legend()
            plt.savefig('n_all.png')

            # m VS avg time
            plt.figure(figsize=(12,8))
            plt.title('number of candidates VS avg time')
            plt.xlabel('Number of candidates')
            plt.ylabel('Avg time (s)')
            plt.grid()
            plt.xticks(candidates)
            avg_times_10 = []
            avg_times_100 = []
            avg_times_200 = []
            avg_times_random = []
            for m in candidates:
                avg_times_10.append(mean([row[3] for row in table if row[1] == m and row[2] == 10]))
                avg_times_100.append(mean([row[3] for row in table if row[1] == m and row[2] == 100]))
                avg_times_200.append(mean([row[3] for row in table if row[1] == m and row[2] == 200]))
                avg_times_random.append(mean([row[2] for row in table_random if row[1] == m]))

            plt.plot(candidates, avg_times_10, marker='o', label='GA: Population size = 10')
            plt.plot(candidates, avg_times_100, marker='o', label='GA: Population size = 100')
            plt.plot(candidates, avg_times_200, marker='o', label='GA: Population size = 200')
            plt.plot(candidates, avg_times_random, marker='o', label='Random generation')

            plt.legend()
            plt.savefig('m_all.png')

            input('\033[96mPress enter to continue...\033[00m')

        elif option == '13':
            # Here, we compare the performance of the genetic algorithm with the random generation for the problem of generating a profile in which a different candidate wins each type of voting rule
            # We are going to generate solutions for:
            # - n in [40,60,80,100,120,140,160,180,200]
            # - m in candidates
            # - population_size = [10, 100, 200]
            # We will generate 10 solutions for each combination of n and m, and plot the average time needed to generate the solution
            # We will do this for both the genetic algorithm and the random generation
            import time
            import matplotlib.pyplot as plt
            import numpy as np
            from tqdm import tqdm
            from statistics import mean

            voters = [50, 100, 200, 300, 500, 1000, 2000]
            candidates = [6, 10, 14, 17, 20]
            pops = [100]

            # Genetic algorithm
            # Table with columns: n, m, population_size, time
            table = []

            for n in tqdm(voters):
                for m in candidates:
                    for population_size in pops:
                        times = []
                        for _ in range(10):
                            start = time.time()
                            _,_ = run_genetic_algorithm(n,m,population_size,mode='diverse')
                            end = time.time()
                            times.append(end-start)
                        table.append([n,m,population_size,mean(times)])

            # Random generation
            table_random = []

            for n in tqdm(voters):
                for m in candidates:
                    times = []
                    for _ in range(10):
                        start = time.time()
                        _,_ = eachWinner(n,m)
                        end = time.time()
                        times.append(end-start)
                    table_random.append([n,m,mean(times)])

            # Plot the results
            # Population size VS avg time
            plt.figure(figsize=(12,8))
            plt.title('Population size VS avg time')
            plt.xlabel('Population size')
            plt.ylabel('Avg time (s)')
            plt.grid()
            plt.xticks(pops)
            avg_times = []
            for population_size in pops:
                avg_times.append(mean([row[3] for row in table if row[2] == population_size]))
            plt.plot(pops, avg_times, marker='o')
            plt.savefig('population_size_diverse.png')

            # n VS avg time
            plt.figure(figsize=(12,8))
            plt.title('number of voters VS avg time')
            plt.xlabel('number of voters')
            plt.ylabel('Avg time (s)')
            plt.grid()
            plt.xticks(voters)
            avg_times_100 = []
            avg_times_random = []
            for n in voters:
                avg_times_100.append(mean([row[3] for row in table if row[0] == n and row[2] == 100]))
                avg_times_random.append(mean([row[2] for row in table_random if row[0] == n]))

            plt.plot(voters, avg_times_100, marker='o', label='GA: Population size = 100')
            plt.plot(voters, avg_times_random, marker='o', label='Random generation')

            plt.legend()
            plt.savefig('n_diverse.png')

            # m VS avg time
            plt.figure(figsize=(12,8))
            plt.title('number of candidates VS avg time')
            plt.xlabel('Number of candidates')
            plt.ylabel('Avg time (s)')
            plt.grid()
            plt.xticks(candidates)
            avg_times_100 = []
            avg_times_random = []

            for m in candidates:
                avg_times_100.append(mean([row[3] for row in table if row[1] == m and row[2] == 100]))
                avg_times_random.append(mean([row[2] for row in table_random if row[1] == m]))

            plt.plot(candidates, avg_times_100, marker='o', label='GA: Population size = 100')
            plt.plot(candidates, avg_times_random, marker='o', label='Random generation')

            plt.legend()
            plt.savefig('m_diverse.png')

        elif option == '14':
            datafile = input('Enter the path to the data file (default: data.xlsx) or input "create" to create a new data file: ')
            if datafile == '':
                datafile = 'data.xlsx'
            if datafile == 'create':
                create_voting_profiles()
            else:
                profiles = read_voting_profiles(datafile)

        elif option == '0':
            print('Bye!')
        else:
            print('\033[91mInvalid option!\033[00m')
            input('\033[96mPress enter to continue...\033[00m')


