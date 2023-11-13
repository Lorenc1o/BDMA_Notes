import openpyxl
import math
import pandas as pd
import random

def excel_to_dict(filename):
    """Converts an Excel file to a dictionary.

    Args:
        filename (str): the name of the Excel file.

    Returns:
        dict: the dictionary corresponding to the Excel file.
            - The keys are the first column of the Excel file.
            - The values are dictionaries:
                - The keys are the column names of the Excel file.
                - The values are the values of the corresponding row.
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(min_row=2):
        data[row[0].value] = {}
        for i in range(1, len(row)): 
            if row[i].value is not None:
                data[row[0].value][ws.cell(row=1, column=i+1).value] = row[i].value
    return data

def invertDict(d):
    """Inverts a dictionary of critiques. That is, the keys are the movies, and the values are dictionaries, where the keys are the persons and the values are the ratings.
    If a person has not rated a movie, we do not add the person to the dictionary.

    Args:
        d (dict): the dictionary to invert.

    Returns:
        dict: the inverted dictionary.
    """
    inverted = {}
    for person in d:
        for movie in d[person]:
            if movie not in inverted:
                inverted[movie] = {}
            inverted[movie][person] = d[person][movie]
    return inverted

def sim_distanceManhattan(person1, person2):
    """Computes the Manhattan distance between two persons.
    
    Args:
        person1 (dict): the first person.
        person2 (dict): the second person.

    Returns:
        float: the Manhattan distance between the two persons.
    """
    distance = 0
    for key in person1:
        if key in person2:
            distance += abs(person1[key] - person2[key])

    return distance

def sim_distanceEuclidean(person1, person2):
    """Computes the Euclidean distance between two persons.
    
    Args:
        person1 (dict): the first person.
        person2 (dict): the second person.

    Returns:
        float: the Euclidean distance between the two persons.
    """
    distance = 0
    for key in person1:
        if key in person2.keys():
            distance += (person1[key] - person2[key])**2

    return distance**0.5

def computeNearestNeighbor(person, critiques, dist="Eucl"):
    """Computes the nearest neighbor of a person.
    
    Args:
        person (dict): the person.
        critiques (dict): the dictionary of critiques.
        dist (str): the distance used to compute the nearest neighbor.

    Returns:
        list: ordered list of the nearest neighbors.
    """
    distances = []
    for critic in critiques:
        if critic != person:
            if dist == "Eucl":
                distances.append((critic, sim_distanceEuclidean(critiques[person], critiques[critic])))
            elif dist == "Manh":
                distances.append((critic, sim_distanceManhattan(critiques[person], critiques[critic])))
    distances.sort(key=lambda x: x[1])
    return distances

def recommendNearestNeighbor(person, critiques, dist="Eucl"):
    """Computes the recommendations for a person based on the nearest neighbor.
    
    Args:
        person (dict): the person.
        critiques (dict): the dictionary of critiques.
        dist (str): the distance used to compute the nearest neighbor.

    Returns:
        list: the list of recommendations for the person.
    """
    recommendations = []
    neighbors = computeNearestNeighbor(person, critiques, dist)
    if len(neighbors) == 0:
        return recommendations
    nearest = neighbors[0][0]
    for movie in critiques[nearest]:
        if movie not in critiques[person]:
            recommendations.append((movie, critiques[nearest][movie]))
    recommendations.sort(key=lambda x: x[1], reverse=True)
    return recommendations

def bestRecommend(person, critiques):
    """Computes the best recommendations for a person.
    
    Args:
        person (dict): the person.
        critiques (dict): the dictionary of critiques.

    Returns:
        list: the list of best recommendations for the person.
    """
    # For each movie not seen by the person, compute the weighted average of the ratings of the nearest neighbors.
    # The weight of a neighbor is the 1/(1 + distance to the person).
    # The rating of a neighbor is the rating of the movie by the neighbor.
    recommendations = []
    inverted = invertDict(critiques)
    for movie in inverted:
        if movie not in critiques[person]:
            total = 0
            s = 0
            for critic in inverted[movie]:
                total += inverted[movie][critic] / (1 + sim_distanceManhattan(critiques[person], critiques[critic]))
                s += 1 / (1 + sim_distanceManhattan(critiques[person], critiques[critic]))

            if s == 0:
                recommendations.append((movie, 0))
            else:
                recommendations.append((movie, total / s))

    recommendations.sort(key=lambda x: x[1], reverse=True)
    return recommendations

def bestRecommentWithExp(person, critiques):
    """Computes the best recommendations for a person.
    
    Args:
        person (dict): the person.
        critiques (dict): the dictionary of critiques.

    Returns:
        list: the list of best recommendations for the person.
    """
    # For each movie not seen by the person, compute the weighted average of the ratings of the nearest neighbors.
    # The weight of a neighbor is the exp(-distance to the person).
    # The rating of a neighbor is the rating of the movie by the neighbor.
    recommendations = []
    inverted = invertDict(critiques)
    for movie in inverted:
        if movie not in critiques[person]:
            total = 0
            s = 0
            for critic in inverted[movie]:
                total += inverted[movie][critic] * math.exp(-sim_distanceManhattan(critiques[person], critiques[critic]))
                s += math.exp(-sim_distanceManhattan(critiques[person], critiques[critic]))

            if s == 0:
                recommendations.append((movie, 0))
            else:
                recommendations.append((movie, total / s))

    recommendations.sort(key=lambda x: x[1], reverse=True)
    return recommendations

def pearson(person1, person2):
    """Computes the Pearson correlation coefficient between two persons.

    Args:
        person1 (dict): the first person.
        person2 (dict): the second person.

    Returns:
        float: the Pearson correlation coefficient between the two persons.
    """
    sum_xy = 0
    sum_x = 0
    sum_y = 0
    sum_x2 = 0
    sum_y2 = 0

    n = 0
    for key in person1:
        if key in person2:
            n += 1
            x = person1[key]
            y = person2[key]
            sum_xy += x * y
            sum_x += x
            sum_y += y
            sum_x2 += x**2
            sum_y2 += y**2

    if n == 0:
        return 0

    denominator = math.sqrt(sum_x2 - (sum_x**2) / n) * math.sqrt(sum_y2 - (sum_y**2) / n)

    if denominator == 0:
        return 0

    return (sum_xy - (sum_x * sum_y) / n) / denominator

def pearsonRecommend(person, critiques):
    """Computes the recommendations for a person based on the Pearson correlation coefficient.
    
    Args:
        person (dict): the person.
        critiques (dict): the dictionary of critiques.

    Returns:
        list: the list of recommendations for the person.
    """
    closest = []
    for critic in critiques:
        if critic != person:
            closest.append((critic, pearson(critiques[person], critiques[critic])))
    closest.sort(key=lambda x: x[1], reverse=True)
    closest = closest[0]

    recommendations = []
    for movie in critiques[closest[0]]:
        if movie not in critiques[person]:
            recommendations.append((movie, critiques[closest[0]][movie]))

    recommendations.sort(key=lambda x: x[1], reverse=True)
    return recommendations, closest[0]

def bestRecommendWithPearson(person, critiques):
    """Computes the best recommendations for a person.
    
    Args:
        person (dict): the person.
        critiques (dict): the dictionary of critiques.

    Returns:
        list: the list of best recommendations for the person.
    """
    # For each movie not seen by the person, compute the weighted average of the ratings of the nearest neighbors.
    # The weight of a neighbor is the exponential of the Pearson correlation coefficient.
    # The rating of a neighbor is the rating of the movie by the neighbor.
    recommendations = []
    inverted = invertDict(critiques)
    for movie in inverted:
        if movie not in critiques[person]:
            total = 0
            s = 0
            for critic in inverted[movie]:
                total += inverted[movie][critic] * math.exp(pearson(critiques[person], critiques[critic]))
                s += math.exp(pearson(critiques[person], critiques[critic]))

            if s == 0:
                recommendations.append((movie, 0))
            else:
                recommendations.append((movie, total / s))

    recommendations.sort(key=lambda x: x[1], reverse=True)
    return recommendations

def cosineCorrelation(person1, person2):
    """ Computes the cosine correlation coefficient between two persons

    Args:
        person1 (dict): the first person
        person2 (dict): the second person

    Returns:
        float: the cosine correlation coefficient between two persons
    """
    sum_xy = 0
    sum_x2 = 0
    sum_y2 = 0

    for key in person1:
        if key in person2:
            x = person1[key]
            y = person2[key]
            sum_xy += x * y
            sum_x2 += x**x
            sum_y2 += y**y
    
    denominator = math.sqrt(sum_x2) * math.sqrt(sum_y2)

    if denominator == 0:
        return 0

    return sum_xy / denominator

def cosineRecommend(person, critiques):
    """Computes the recommendations for a person based on the cosine correlation coefficient.
    
    Args:
        person (dict): the person.
        critiques (dict): the dictionary of critiques.

    Returns:
        list: the list of recommendations for the person.
    """
    closest = []
    for critic in critiques:
        if critic != person:
            closest.append((critic, cosineCorrelation(critiques[person], critiques[critic])))
    closest.sort(key=lambda x: x[1], reverse=True)
    closest = closest[0]

    recommendations = []
    for movie in critiques[closest[0]]:
        if movie not in critiques[person]:
            recommendations.append((movie, critiques[closest[0]][movie]))

    recommendations.sort(key=lambda x: x[1], reverse=True)
    return recommendations, closest[0]

def bestRecommendWithCosine(person, critiques):
    """Computes the best recommendations for a person.
    
    Args:
        person (dict): the person.
        critiques (dict): the dictionary of critiques.

    Returns:
        list: the list of best recommendations for the person.
    """
    # For each movie not seen by the person, compute the weighted average of the ratings of the nearest neighbors.
    # The weight of a neighbor is the exponential of the cosine correlation coefficient.
    # The rating of a neighbor is the rating of the movie by the neighbor.
    recommendations = []
    inverted = invertDict(critiques)
    for movie in inverted:
        if movie not in critiques[person]:
            total = 0
            s = 0
            for critic in inverted[movie]:
                total += inverted[movie][critic] * math.exp(cosineCorrelation(critiques[person], critiques[critic]))
                s += math.exp(cosineCorrelation(critiques[person], critiques[critic]))

            if s == 0:
                recommendations.append((movie, 0))
            else:
                recommendations.append((movie, total / s))

    recommendations.sort(key=lambda x: x[1], reverse=True)
    return recommendations

def getTableOfRecommendations(users, critiques):
    """For each user, computes the top 1 recommendations using the different methods.
    
    Args:
        users (list): the list of users.
        critiques (dict): the dictionary of critiques.

    Returns:
        pd.DataFrame: the table of recommendations.
    """
    table = pd.DataFrame(columns=["Nearest neighbor", "Best", "Best with exponential function", "Pearson", "Best with Pearson", "Cosine", "Best with cosine"])
    for user in users:
        table.loc[user] = [recommendNearestNeighbor(user, critiques, 'Manh')[0][0], bestRecommend(user, critiques)[0][0], bestRecommentWithExp(user, critiques)[0][0], pearsonRecommend(user, critiques)[0][0][0], bestRecommendWithPearson(user, critiques)[0][0], cosineRecommend(user, critiques)[0][0][0], bestRecommendWithCosine(user, critiques)[0][0]]
    return table

def getTableOfRecommendations_v2(users, critiques):
    """For each user, computes the top 1 recommendations using the different methods.
    
    Args:
        users (list): the list of users.
        critiques (dict): the dictionary of critiques.

    Returns:
        pd.DataFrame: the table of recommendations.
    """
    table = pd.DataFrame(columns=["Nearest neighbor", "Best", "Best with exponential function", "Best with Pearson", "Best with cosine"])
    for user in users:
        table.loc[user] = [recommendNearestNeighbor(user, critiques, 'Manh')[0][0], bestRecommend(user, critiques)[0][0], bestRecommentWithExp(user, critiques)[0][0], bestRecommendWithPearson(user, critiques)[0][0], bestRecommendWithCosine(user, critiques)[0][0]]
    return table

def checkMinBlanks(critiques):
    """Checks if the table of critiques has at least 30% of blanks.
    
    Args:
        critiques (dict): the dictionary of critiques.

    Returns:
        bool: True if the table of critiques has at least 30% of blanks, False otherwise.
    """
    nbBlanks = 0
    for person in critiques:
        for movie in critiques[person]:
            if critiques[person][movie] == "":
                nbBlanks += 1
    return nbBlanks >= 0.3 * len(critiques) * len(critiques[list(critiques.keys())[0]])

def checkMaxBlanks(critiques):
    """Checks if the table of critiques has at most 50% of blanks.
    
    Args:
        critiques (dict): the dictionary of critiques.

    Returns:
        bool: True if the table of critiques has at most 70% of blanks, False otherwise.
    """
    nbBlanks = 0
    for person in critiques:
        for movie in critiques[person]:
            if critiques[person][movie] == "":
                nbBlanks += 1
    return nbBlanks <= 0.5 * len(critiques) * len(critiques[list(critiques.keys())[0]])

def findUser(critiques):
    """Finds a user that has not rated more than 50% of the movies.
    
    Args:
        critiques (dict): the dictionary of critiques.

    Returns:
        str: the name of the user.
    """
    # First, check if there is such a user
    n_movies = len(critiques[list(critiques.keys())[0]])
    for user in critiques:
        nbBlanks = 0
        for movie in critiques[user]:
            if critiques[user][movie] == "":
                nbBlanks += 1
        if nbBlanks >= 0.5 * n_movies and nbBlanks >= 2: # We want at least 2 rated movies to be able to compute the recommendations
            return user
    return None

def CritiquesGenerator(n = 10, m = 15):
    """Generates a table of critiques with n persons and m movies.
    
    Args:
        n (int): the number of persons.
        m (int): the number of movies.

    Returns:
        dict: the dictionary of critiques.
    """
    critiques = {}
    for i in range(n):
        critiques["Person "+str(i+1)] = {}
        for j in range(m):
            if random.random() < 0.2:
                critiques["Person "+str(i+1)]["Movie "+str(j+1)] = ""
            else:
                critiques["Person "+str(i+1)]["Movie "+str(j+1)] = random.randint(1, 5)
    return critiques

def ValidCritiquesGenerator(n=10, m=15):
    """Generates a table of critiques with n persons and m movies that has at least 30% of blanks and at most 50% of blanks.
    
    Args:
        n (int): the number of persons.
        m (int): the number of movies.

    Returns:
        dict: the dictionary of critiques.
    """
    critiques = CritiquesGenerator(n, m)
    user = None
    while not checkMinBlanks(critiques) or not checkMaxBlanks(critiques) or user is None:
        critiques = CritiquesGenerator(n, m)
        user = findUser(critiques)

    # Once we have it, we remove the blanks
    final = {}

    for person in critiques:
        final[person] = {}
        for movie in critiques[person]:
            if critiques[person][movie] != "":
                final[person][movie] = critiques[person][movie]
    
    return final, user

def AllRecommend(n=10, m=15, n_iter=None):
    """Generates a table of critiques with n persons and m movies and finds a user that has not rated more than 50% of the movies.
        In this case, the user receives the same recommendations for all the methods.

    Args:
        n (int): the number of persons.
        m (int): the number of movies.

    Returns:
        dict: the dictionary of critiques.
    """
    k = 0
    while (n_iter is None and True) or (k < n_iter):
        critiques, user = ValidCritiquesGenerator(n, m)

        # Get the recommendation table
        table = getTableOfRecommendations_v2([user], critiques)

        # Check if the recommendations are the same for all the methods (there is only one movie)
        if len(table.iloc[0].unique()) == 1:
            return critiques, user, table
        
        k += 1

    return None, None, None

def DiffRecommend(n=10, m=15, n_iter=None):
    """Generates a table of critiques with n persons and m movies and finds a user that has not rated more than 50% of the movies.
        In this case, the user receives different recommendations for the different methods.

    Args:
        n (int): the number of persons.
        m (int): the number of movies.

    Returns:
        dict: the dictionary of critiques.
    """
    k = 0
    while (n_iter is None and True) or (k < n_iter):
        critiques, user = ValidCritiquesGenerator(n, m)

        # Get the recommendation table
        table = getTableOfRecommendations_v2([user], critiques)

        # Check if the recommendations are the same for all the methods (there is only one movie)
        if len(table.iloc[0].unique()) == 5:
            return critiques, user, table
        
        k += 1

    return None, None, None

if __name__ == "__main__":
    filename = "data.xlsx"

    # Exercise 1: Convert the Excel file to a dictionary.
    print("\033[93mReading Excel file...")
    critiques = excel_to_dict(filename)
    print("Done.\033[0m")
    print(critiques["Lisa Rose"])

    # Exercise 2: Find similar persons and compute recommendations.
    # a) Compute the Manhattan and Euclidean distances between Lisa Rose and Gene Seymour.
    print("\033[93mComputing Manhattan distance...\033[0m")
    print(sim_distanceManhattan(critiques["Lisa Rose"], critiques["Gene Seymour"]))
    print("\033[93mComputing Euclidean distance...\033[0m")
    print(sim_distanceEuclidean(critiques["Lisa Rose"], critiques["Gene Seymour"]))

    # b) Get recommendations based on nearest neighbors using Manhattan distance.
    print("\033[93mComputing nearest neighbors...\033[0m")
    print("\033[94mFor Lisa Rose:\033[0m")
    print(computeNearestNeighbor("Lisa Rose", critiques, "Manh"))
    print("\033[94mFor Toby:\033[0m")
    print(computeNearestNeighbor("Toby", critiques, "Manh"))
    print("\033[93mComputing recommendations...\033[0m")
    print("\033[94mFor Lisa Rose:\033[0m")
    print(recommendNearestNeighbor("Lisa Rose", critiques, "Manh"))
    print("\033[94mFor Toby:\033[0m")
    print(recommendNearestNeighbor("Toby", critiques, "Manh"))

    # c) 
    #   i) Rank the critics and recommend the best movies by using the weighted average of the ratings.
    print("\033[93mComputing best recommendations...\033[0m")
    print("\033[94mFor Anne:\033[0m")
    print(bestRecommend("Anne", critiques))

    #   ii) Rank the critics and recommend the best movies by using the weighted average of the ratings and the exponential function.
    print("\033[93mComputing best recommendations with exponential function...\033[0m")
    print("\033[94mFor Anne:\033[0m")
    print(bestRecommentWithExp("Anne", critiques))

    # d) Use the Pearson correlation coefficient to get recommendations
    print("\033[93mComputing Pearson recommendations...\033[0m")
    print("For Anne, the pearson correlations are:")
    for critic in critiques:
        if critic != "Anne":
            print("\t"+critic+": "+str(pearson(critiques["Anne"], critiques[critic])))
    print("For Anne:, the recommendations are:")
    print(pearsonRecommend("Anne", critiques))
    print("If we weight the recommendations by the Pearson correlation coefficient, we get:")
    print(bestRecommendWithPearson("Anne", critiques))

    # e) Use the cosine correlation coefficient to get recommendations
    print("\033[93mComputing cosine recommendations...\033[0m")
    print("\033[94mFor Anne:\033[0m")
    print(cosineRecommend("Anne", critiques))
    print("If we weight the recommendations by the cosine correlation coefficient, we get:")
    print(bestRecommendWithCosine("Anne", critiques))

    # Exercise 3: Use the other dataset. data1.xlsx
    print("\033[93mReading Excel file...")
    critiques1 = excel_to_dict("data1.xlsx")
    print("Done.\033[0m")
    for person in critiques1:
        print(person+": "+str(critiques1[person]))

    # Get all different recommendations for Veronica and Hailey
    print("\033[93mComputing recommendations...\033[0m")
    print("\033[94mFor Veronica:\033[0m")
    print("\tNearest neighbor:")
    print("\t\t"+str(recommendNearestNeighbor("Veronica", critiques1)))
    print("\tBest:")
    print("\t\t"+str(bestRecommend("Veronica", critiques1)))
    print("\tBest with exponential function:")  
    print("\t\t"+str(bestRecommentWithExp("Veronica", critiques1)))
    print("\tPearson:")
    print("\t\t"+str(pearsonRecommend("Veronica", critiques1)))
    print("\tBest with Pearson:")
    print("\t\t"+str(bestRecommendWithPearson("Veronica", critiques1)))
    print("\tCosine:")
    print("\t\t"+str(cosineRecommend("Veronica", critiques1)))
    print("\tBest with cosine:")
    print("\t\t"+str(bestRecommendWithCosine("Veronica", critiques1)))

    print("\033[94mFor Hailey:\033[0m")
    print("\tNearest neighbor:")
    print("\t\t"+str(recommendNearestNeighbor("Hailey", critiques1)))
    print("\tBest:")
    print("\t\t"+str(bestRecommend("Hailey", critiques1)))
    print("\tBest with exponential function:")
    print("\t\t"+str(bestRecommentWithExp("Hailey", critiques1)))
    print("\tPearson:")
    print("\t\t"+str(pearsonRecommend("Hailey", critiques1)))
    print("\tBest with Pearson:")
    print("\t\t"+str(bestRecommendWithPearson("Hailey", critiques1)))
    print("\tCosine:")
    print("\t\t"+str(cosineRecommend("Hailey", critiques1)))
    print("\tBest with cosine:")
    print("\t\t"+str(bestRecommendWithCosine("Hailey", critiques1)))

    # Get the table of recommendations for Anne
    print("\033[93mComputing table of recommendations...\033[0m")
    print(getTableOfRecommendations(["Anne"], critiques))

    # Get the table of recommendations for Veronica and Hailey
    print("\033[93mComputing table of recommendations...\033[0m")
    print(getTableOfRecommendations(["Veronica", "Hailey"], critiques1))

    # Exercise 4: Generate a table of critiques with n persons and m movies such that:
    #   - The table has at least 30% of blanks.
    #   - The table has at most 50% of blanks.
    #   - There is a user that has not rated more than 50% of the movies.
    #   - The recommendations are the same for all the methods.
    n = 10
    m = 15
    print("\033[93mGenerating table of critiques...\033[0m")
    critiques, user, table = AllRecommend(n, m)
    print("\033[94mThe table of critiques is:\033[0m")
    # Reformat from dict to pd.DataFrame
    df = pd.DataFrame.from_dict(critiques)
    print(df)
    n_nans = df.isna().sum().sum()
    print("There are "+str(n_nans)+" NaNs in the table, that is "+str(n_nans / (n*m) * 100)+"% of the table.")
    print("\033[94mThe table of recommendations is:\033[0m")
    print(table)

    # Exercise 5: Generate a table of critiques with n persons and m movies such that:
    #   - The table has at least 30% of blanks.
    #   - The table has at most 50% of blanks.
    #   - There is a user that has not rated more than 50% of the movies.
    #   - The recommendations are different for the different methods.
    print("\033[93mGenerating table of critiques...\033[0m")
    critiques, user, table = DiffRecommend(10, 15)
    print("\033[94mThe table of critiques is:\033[0m")
    # Reformat from dict to pd.DataFrame
    df = pd.DataFrame.from_dict(critiques)
    print(df)
    n_nans = df.isna().sum().sum()
    print("There are "+str(n_nans)+" NaNs in the table, that is "+str(n_nans / (n*m) * 100)+"% of the table.")
    print("\033[94mThe table of recommendations is:\033[0m")
    print(table)








